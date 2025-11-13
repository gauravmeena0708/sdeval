#!/usr/bin/env python3
"""Bulk-evaluate synthetic CSVs, summarize metrics, and generate visuals."""

from __future__ import annotations

import argparse
import json
import math
import subprocess
import sys
from pathlib import Path
from typing import Any, Dict, Iterable, List, Sequence, Tuple


# Metrics where lower values are preferred.
LOWER_IS_BETTER = {
    "statistical.mean_abs_mean_diff",
    "statistical.mean_abs_std_diff",
    "statistical.avg_wasserstein",
    "coverage.missing_category_ratio",
    "coverage.missingness_delta",
    "privacy.dcr",
    "privacy.knn_distance",
}


VISUALIZATION_GROUPS = [
    (
        "Statistical",
        [
            ("statistical.alpha_precision", "Alpha Precision", "direct"),
            ("statistical.beta_recall", "Beta Recall", "direct"),
        ],
    ),
    (
        "Coverage",
        [
            ("coverage.unique_ratio", "Unique Ratio", "direct"),
            ("coverage.rare_category_retention", "Rare Retention", "direct"),
            ("coverage.missing_category_ratio", "1 - Missing Ratio", "invert"),
        ],
    ),
    (
        "Privacy",
        [
            ("privacy.dcr", "1 - DCR", "invert"),
            ("privacy.nndr", "1 - NNDR", "invert"),
            ("privacy.knn_distance", "KNN Distance", "knn"),
        ],
    ),
]

RADAR_METRICS = {
    "data_quality": [
        ("statistical.alpha_precision", "Alpha Precision", "direct"),
        ("statistical.beta_recall", "Beta Recall", "direct"),
        ("coverage.unique_ratio", "Unique Ratio", "direct"),
        ("coverage.rare_category_retention", "Rare Retention", "direct"),
    ],
    "privacy": [
        ("privacy.knn_distance", "KNN Distance", "knn"),
        ("privacy.dcr", "1 - DCR", "invert"),
        ("privacy.nndr", "1 - NNDR", "invert"),
    ],
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--synthetic-dir",
        required=True,
        help="Directory containing synthetic CSV files.",
    )
    parser.add_argument(
        "--synthetic-pattern",
        default="*.csv",
        help="Glob pattern to select synthetic CSVs (default: *.csv).",
    )
    parser.add_argument(
        "--real-data-csv",
        required=True,
        help="Path to the real data CSV used for evaluation.",
    )
    parser.add_argument(
        "--output-dir",
        default="outputs",
        help="Directory to store evaluation outputs (default: outputs).",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Pass --overwrite to sdeval to replace existing outputs.",
    )
    parser.add_argument(
        "--metrics",
        nargs="*",
        help="Optional subset of metrics to display (e.g. statistical.alpha_precision).",
    )
    parser.add_argument(
        "--show-non-numeric",
        action="store_true",
        help="Include non-numeric metrics (booleans/strings) in the table/Excel.",
    )
    parser.add_argument(
        "--excel-path",
        default=None,
        help="Path to write Excel summary (default: <output-dir>/summary.xlsx).",
    )
    parser.add_argument(
        "--visualization-dir",
        default=None,
        help="Directory for PNG dashboards (default: <output-dir>/visualizations).",
    )
    parser.add_argument(
        "--visualize",
        action="store_true",
        help="Enable visual outputs (dashboards plus per-file diagnostics).",
    )
    return parser.parse_args()


def find_synthetic_csvs(directory: Path, pattern: str) -> List[Path]:
    return sorted(directory.glob(pattern))


def run_sdeval(csv_path: Path, real_data_csv: Path, output_dir: Path, overwrite: bool, visualize: bool) -> None:
    cmd = [
        sys.executable,
        "-m",
        "sdeval.main",
        "--input-path",
        str(csv_path),
        "--real-data-csv-path",
        str(real_data_csv),
        "--output-dir",
        str(output_dir),
    ]
    if overwrite:
        cmd.append("--overwrite")
    if visualize:
        cmd.append("--visualize")
    print(f"[RUN] Evaluating {csv_path.name}")
    subprocess.run(cmd, check=True)


def _dedupe_parts(parts: Sequence[str]) -> List[str]:
    result: List[str] = []
    for part in parts:
        if result and part.startswith(result[-1] + "_"):
            trimmed = part[len(result[-1]) + 1 :]
            result.append(trimmed)
        else:
            result.append(part)
    return result


def flatten_metrics(data: Dict[str, Any], prefix: Tuple[str, ...] = ()) -> Dict[str, Any]:
    flat: Dict[str, Any] = {}
    for key, value in data.items():
        new_prefix = prefix + (key,)
        if isinstance(value, dict):
            flat.update(flatten_metrics(value, new_prefix))
        else:
            pretty_key = ".".join(_dedupe_parts(new_prefix))
            flat[pretty_key] = value
    return flat


def load_summary(path: Path) -> Dict[str, Any]:
    with path.open("r", encoding="utf-8") as f:
        data = json.load(f)
    flat = flatten_metrics(data)
    flat["file"] = path.name
    return flat


def collect_metrics(rows: List[Dict[str, Any]], requested: Iterable[str] | None, include_non_numeric: bool) -> List[str]:
    all_keys = set().union(*(row.keys() for row in rows))
    all_keys.discard("file")
    if requested:
        normalized = []
        for metric in requested:
            if metric not in all_keys:
                raise ValueError(f"Metric '{metric}' not found in summaries.")
            normalized.append(metric)
        return normalized
    metrics = []
    for key in sorted(all_keys):
        sample_value = next((row.get(key) for row in rows if key in row), None)
        if include_non_numeric:
            metrics.append(key)
        else:
            if isinstance(sample_value, bool) or sample_value is None:
                continue
            if isinstance(sample_value, (int, float)):
                metrics.append(key)
    return metrics


def metric_orientation(metric: str) -> str:
    return "lower" if metric in LOWER_IS_BETTER else "higher"


def compute_highlights(rows: List[Dict[str, Any]], metrics: Sequence[str]) -> Dict[Tuple[int, str], str]:
    highlights: Dict[Tuple[int, str], str] = {}
    for metric in metrics:
        values: List[Tuple[int, float]] = []
        for idx, row in enumerate(rows):
            value = row.get(metric)
            if value is None or isinstance(value, bool):
                continue
            if isinstance(value, (int, float)):
                num_value = float(value)
                if math.isnan(num_value):
                    continue
                values.append((idx, num_value))
        if not values:
            continue
        orient = metric_orientation(metric)
        best_value = min(v for _, v in values) if orient == "lower" else max(v for _, v in values)
        worst_value = max(v for _, v in values) if orient == "lower" else min(v for _, v in values)
        if math.isclose(best_value, worst_value, rel_tol=1e-9, abs_tol=1e-12):
            continue
        for idx, val in values:
            if math.isclose(val, best_value, rel_tol=1e-9, abs_tol=1e-12):
                highlights[(idx, metric)] = "best"
            elif math.isclose(val, worst_value, rel_tol=1e-9, abs_tol=1e-12):
                highlights[(idx, metric)] = "worst"
    return highlights


def format_value(value: Any) -> str:
    if value is None:
        return "-"
    if isinstance(value, bool):
        return "yes" if value else "no"
    if isinstance(value, (int, float)):
        return f"{value:.6g}"
    return str(value)


def render_table(rows: List[Dict[str, Any]], metrics: Sequence[str], highlights: Dict[Tuple[int, str], str]) -> str:
    columns = ["file", *metrics]
    formatted_rows: List[Dict[str, str]] = []
    for idx, row in enumerate(rows):
        formatted: Dict[str, str] = {"file": row["file"]}
        for metric in metrics:
            value = row.get(metric)
            text = format_value(value)
            marker = highlights.get((idx, metric))
            if marker == "best":
                text = f"{text} *"
            elif marker == "worst":
                text = f"{text} !"
            formatted[metric] = text
        formatted_rows.append(formatted)

    widths = {col: len(col) for col in columns}
    for row in formatted_rows:
        for col in columns:
            widths[col] = max(widths[col], len(row.get(col, "")))

    lines = []
    header = " | ".join(col.ljust(widths[col]) for col in columns)
    divider = "-+-".join("-" * widths[col] for col in columns)
    lines.append(header)
    lines.append(divider)
    for row in formatted_rows:
        line = " | ".join(row.get(col, "").ljust(widths[col]) for col in columns)
        lines.append(line)
    lines.append("\nLegend: '*' best value, '!' worst value, '-' missing.")
    return "\n".join(lines)


def build_dataframe(rows: List[Dict[str, Any]], metrics: Sequence[str]):
    import pandas as pd  # type: ignore

    data = []
    for row in rows:
        entry = {"file": row["file"]}
        for metric in metrics:
            entry[metric] = row.get(metric)
        data.append(entry)
    return pd.DataFrame(data)


def write_excel(rows: List[Dict[str, Any]], metrics: Sequence[str], highlights: Dict[Tuple[int, str], str], excel_path: Path) -> None:
    import pandas as pd  # type: ignore
    from openpyxl.styles import PatternFill  # type: ignore
    from openpyxl.utils import get_column_letter  # type: ignore

    df = build_dataframe(rows, metrics)
    excel_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="metrics")
        sheet = writer.sheets["metrics"]

        best_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        worst_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")

        for (row_idx, metric), tag in highlights.items():
            if metric not in df.columns:
                continue
            excel_row = 2 + row_idx
            excel_col = df.columns.get_loc(metric) + 1
            cell = sheet.cell(row=excel_row, column=excel_col)
            if tag == "best":
                cell.fill = best_fill
            elif tag == "worst":
                cell.fill = worst_fill

        for idx, column in enumerate(df.columns, start=1):
            max_len = max(
                [len(str(column))] + [len(str(val)) if val is not None else 1 for val in df[column]]
            )
            sheet.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 60)

        sheet.freeze_panes = "B2"


def clamp01(value: float) -> float:
    return max(0.0, min(1.0, value))


def normalize_for_visual(raw_value: Any, mode: str, context: Dict[str, float]) -> float:
    if raw_value is None:
        return 0.0
    try:
        value = float(raw_value)
    except (TypeError, ValueError):
        return 0.0

    if mode == "direct":
        return clamp01(value)
    if mode == "invert":
        return clamp01(1.0 - value)
    if mode == "knn":
        max_knn = context.get("max_knn_distance", 1.0) or 1.0
        return clamp01(value / max_knn)
    return 0.0


def generate_visualizations(rows: List[Dict[str, Any]], vis_dir: Path) -> List[Path]:
    try:
        import matplotlib.pyplot as plt  # type: ignore
        import numpy as np  # type: ignore
    except ImportError as exc:
        raise SystemExit("matplotlib is required for visualization output.") from exc

    vis_dir.mkdir(parents=True, exist_ok=True)
    cmap = plt.get_cmap("RdYlGn")
    max_knn = max(
        (float(row.get("privacy.knn_distance") or 0.0) for row in rows),
        default=0.0,
    )
    context = {"max_knn_distance": max_knn if max_knn > 0 else 1.0}

    generated_paths: List[Path] = []
    for row in rows:
        file_name = row["file"]
        title = file_name.replace("_summary.json", "")
        fig, axes = plt.subplots(1, len(VISUALIZATION_GROUPS), figsize=(4 * len(VISUALIZATION_GROUPS), 4), sharey=True)
        if len(VISUALIZATION_GROUPS) == 1:
            axes = [axes]

        for ax, (group_name, metrics) in zip(axes, VISUALIZATION_GROUPS):
            labels: List[str] = []
            normalized: List[float] = []
            actuals: List[str] = []
            for metric_key, label, mode in metrics:
                labels.append(label)
                normalized.append(normalize_for_visual(row.get(metric_key), mode, context))
                val = row.get(metric_key)
                if isinstance(val, (int, float)):
                    actuals.append(f"{val:.3g}")
                elif val is None:
                    actuals.append("-")
                else:
                    actuals.append(str(val))
            bar_colors = [cmap(value) for value in normalized]
            bars = ax.bar(range(len(labels)), normalized, color=bar_colors)
            ax.set_xticks(range(len(labels)))
            ax.set_xticklabels(labels, rotation=40, ha="right")
            ax.set_ylim(0, 1.05)
            ax.set_title(group_name)
            for bar, text in zip(bars, actuals):
                ax.text(
                    bar.get_x() + bar.get_width() / 2,
                    bar.get_height() + 0.02,
                    text,
                    ha="center",
                    va="bottom",
                    fontsize=8,
                )
        fig.suptitle(f"Quality overview: {title}", fontsize=14)
        fig.tight_layout(rect=[0, 0, 1, 0.95])
        subdir = vis_dir / title
        subdir.mkdir(parents=True, exist_ok=True)
        quality_path = subdir / f"{title}_quality.png"
        fig.savefig(quality_path, dpi=200)
        plt.close(fig)
        generated_paths.append(quality_path)

        for radar_type, metrics in RADAR_METRICS.items():
            radar_labels = [label for _, label, _ in metrics]
            radar_values = [normalize_for_visual(row.get(metric), mode, context) for metric, _, mode in metrics]
            if not any(radar_values):
                continue
            angles = np.linspace(0, 2 * np.pi, len(radar_labels), endpoint=False)
            angles = np.concatenate([angles, angles[:1]])
            values = radar_values + radar_values[:1]
            fig, ax = plt.subplots(subplot_kw={"projection": "polar"})
            ax.plot(angles, values, color="#4c78a8", linewidth=2)
            ax.fill(angles, values, color="#4c78a8", alpha=0.3)
            ax.set_xticks(angles[:-1])
            ax.set_xticklabels(radar_labels, fontsize=8)
            ax.set_yticklabels([])
            ax.set_ylim(0, 1)
            pretty_name = "Data Quality" if radar_type == "data_quality" else "Privacy"
            ax.set_title(f"{pretty_name} Radar: {title}", fontsize=12)
            radar_path = subdir / f"{title}_{radar_type}.png"
            fig.tight_layout()
            fig.savefig(radar_path, dpi=200)
            plt.close(fig)
            generated_paths.append(radar_path)

    return generated_paths


def main() -> None:
    args = parse_args()
    synthetic_dir = Path(args.synthetic_dir)
    if not synthetic_dir.exists():
        raise SystemExit(f"Synthetic directory not found: {synthetic_dir}")

    real_data_csv = Path(args.real_data_csv)
    if not real_data_csv.exists():
        raise SystemExit(f"Real data CSV not found: {real_data_csv}")

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    csv_paths = find_synthetic_csvs(synthetic_dir, args.synthetic_pattern)
    if not csv_paths:
        raise SystemExit(f"No CSV files found in {synthetic_dir} matching {args.synthetic_pattern}")

    for csv_path in csv_paths:
        run_sdeval(csv_path, real_data_csv, output_dir, args.overwrite, args.visualize)

    summary_paths: List[Path] = []
    for csv_path in csv_paths:
        summary_path = output_dir / f"{csv_path.stem}_summary.json"
        if summary_path.exists():
            summary_paths.append(summary_path)
        else:
            print(f"[WARN] Summary file missing for {csv_path.name}: {summary_path}")

    if not summary_paths:
        raise SystemExit("No summary JSON files were created; aborting.")

    rows = [load_summary(path) for path in summary_paths]
    metrics = collect_metrics(rows, args.metrics, args.show_non_numeric)
    if not metrics:
        raise SystemExit("No metrics available to display with the current options.")

    highlights = compute_highlights(rows, metrics)
    table = render_table(rows, metrics, highlights)
    print("\n" + table)

    excel_path = Path(args.excel_path) if args.excel_path else output_dir / "summary.xlsx"
    write_excel(rows, metrics, highlights, excel_path)
    print(f"\nExcel report saved to {excel_path}")

    if args.visualize:
        vis_dir = Path(args.visualization_dir) if args.visualization_dir else output_dir / "visualizations"
        visuals = generate_visualizations(rows, vis_dir)
        if visuals:
            print("Visualization files:")
            for path in visuals:
                print(f"  - {path}")
        print("Detailed per-file diagnostics are stored inside the same visualizations directory.")


if __name__ == "__main__":
    main()
